"""
Buyer Match — on-demand mandate pipeline (Data Engine).

Faithful port of match_server.py's load_mandate + doc summarisers: build the mandate query
text from DB fields, then fetch + parse each document (PDF via pdftotext/PyPDF2, Excel via
openpyxl), gpt-4o-mini summarise, and assemble the `full_text` that gets embedded.

Reads mandate metadata from buyer_match.mandates (the synced replica). Document files come
from `docs_base` — local disk (dev) or an SFTP-mirrored path (prod, Phase 5).
"""
import hashlib
import json
import os
import subprocess

import httpx
import psycopg2.extras

DOCS_BASE_DEFAULT = ("/Users/craiganderson/Dropbox/dev/on-testing/"
                     "origryxd-2026-06-21/origryxd/homedir/public_html/library/storage/app/public")
GPT_MODEL = "gpt-4o-mini"
PDF_PROMPT = ("Summarise this M&A document in 3-5 bullet points. Focus on: what the company "
              "does, key financials, sectors served, and any unique selling points. Be concise.")
XLS_PROMPT = ("Summarise this M&A financial data in 3-5 bullet points. Focus on: revenue, "
              "EBITDA, growth trends, and any notable metrics. Be concise.")


def _key():
    k = os.environ.get("OPENAI_API_KEY")
    if not k:
        raise RuntimeError("OPENAI_API_KEY not set")
    return k


def _call_gpt_summarise(text, system_prompt):
    r = httpx.post(
        "https://api.openai.com/v1/chat/completions",
        headers={"Authorization": f"Bearer {_key()}", "Content-Type": "application/json"},
        json={"model": GPT_MODEL, "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": text}],
            "max_tokens": 300, "temperature": 0.3},
        timeout=60,
    )
    r.raise_for_status()
    d = r.json()
    return d["choices"][0]["message"]["content"], d.get("usage", {})


def _summarise_pdf(path):
    try:
        try:
            res = subprocess.run(["pdftotext", "-layout", path, "-"],
                                 capture_output=True, text=True, timeout=30)
            text = res.stdout.strip()
        except FileNotFoundError:
            try:
                try:
                    from pypdf import PdfReader
                except ImportError:
                    from PyPDF2 import PdfReader
                text = "\n".join(p.extract_text() or "" for p in PdfReader(path).pages)
            except ImportError:
                return "(PDF extraction not available — install pdftotext or pypdf)", {}
        if not text or len(text) < 20:
            return "(PDF contained no extractable text)", {}
        if len(text) > 8000:
            text = text[:8000] + "..."
        return _call_gpt_summarise(text, PDF_PROMPT)
    except Exception as e:  # noqa: BLE001
        return f"(Error summarising PDF: {e})", {}


def _summarise_excel(path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames[:5]:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(max_row=50, values_only=True):
                rt = " | ".join(str(c) for c in row if c is not None)
                if rt.strip():
                    rows.append(rt)
            if rows:
                parts.append(f"Sheet: {sheet}\n" + "\n".join(rows))
        wb.close()
        text = "\n\n".join(parts)
        if not text or len(text) < 20:
            return "(Excel contained no extractable data)", {}
        if len(text) > 8000:
            text = text[:8000] + "..."
        return _call_gpt_summarise(text, XLS_PROMPT)
    except ImportError:
        return "(Excel extraction not available — install openpyxl)", {}
    except Exception as e:  # noqa: BLE001
        return f"(Error summarising Excel: {e})", {}


def _as_list(v):
    if v is None:
        return []
    if isinstance(v, str):
        try:
            v = json.loads(v)
        except (json.JSONDecodeError, TypeError):
            return []
    return v if isinstance(v, list) else []


def build_mandate_text(m):
    """EXACT parity with match_server.load_mandate's mandate_text assembly."""
    parts = []
    if m.get("summary"):
        parts.append(f"Summary: {m['summary']}")
    if m.get("points_paragraph_top"):
        parts.append(f"\n{m['points_paragraph_top']}")
    pts = _as_list(m.get("points"))
    if pts:
        bullets = "\n".join(f"• {p.get('text', p) if isinstance(p, dict) else p}" for p in pts)
        parts.append(f"\n{bullets}")
    return "\n".join(parts)


def _doc_hash(path):
    return hashlib.sha256((GPT_MODEL + "\n" + (path or "")).encode("utf-8")).hexdigest()


def _sftp_fetch(doc_path):
    """PROD: pull a mandate document from the ON storage server over SFTP to a temp file.
    Returns the temp path (caller deletes) or None if SFTP isn't configured / the fetch fails.
    Only hit on a doc_cache MISS, so at most once per document."""
    host = os.environ.get("BM_SFTP_HOST")
    if not host or not doc_path:
        return None
    import tempfile
    try:
        import paramiko
    except ImportError:
        return None
    base = (os.environ.get("BM_SFTP_BASE", "") or "").rstrip("/")
    remote = base + "/" + doc_path.lstrip("/")
    fd, tmp = tempfile.mkstemp(suffix=os.path.splitext(doc_path)[1] or ".bin")
    os.close(fd)
    cli = paramiko.SSHClient()
    cli.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        cli.connect(host, port=int(os.environ.get("BM_SFTP_PORT", "22")),
                    username=os.environ.get("BM_SFTP_USER"),
                    key_filename=os.environ.get("BM_SFTP_KEY"), timeout=25)
        sftp = cli.open_sftp()
        try:
            sftp.get(remote, tmp)
        finally:
            sftp.close()
        return tmp
    except Exception:  # noqa: BLE001 — missing file / auth / network → treat as not found
        try:
            os.remove(tmp)
        except OSError:
            pass
        return None
    finally:
        cli.close()


def load_mandate(conn, identifier, docs_base=None):
    """Fetch a mandate (by code, else numeric id) and run the doc pipeline. Mirrors the tool."""
    docs_base = docs_base or os.environ.get("BM_DOCS_BASE", DOCS_BASE_DEFAULT)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        m = None
        for tbl in ("opportunities", "bs_opportunities"):   # opportunities first, like the tool
            cur.execute("SELECT * FROM buyer_match.mandates WHERE source_table=%s AND code=%s", (tbl, identifier))
            m = cur.fetchone()
            if not m and str(identifier).isdigit():
                cur.execute("SELECT * FROM buyer_match.mandates WHERE source_table=%s AND id=%s", (tbl, int(identifier)))
                m = cur.fetchone()
            if m:
                break
    if not m:
        return {"error": f"Mandate '{identifier}' not found"}

    mandate_text = build_mandate_text(m)

    doc_summaries, ptok, ctok = [], 0, 0
    for doc in _as_list(m.get("documents")):
        title = doc.get("title", "Untitled")
        doc_path = doc.get("document", "")
        dh = _doc_hash(doc_path)
        # cache hit -> reuse summary, skip fetch + gpt (the cost)
        with conn.cursor() as cur:
            cur.execute("SELECT summary FROM buyer_match.doc_cache WHERE doc_hash=%s", (dh,))
            crow = cur.fetchone()
        if crow:
            with conn.cursor() as cur:
                cur.execute("UPDATE buyer_match.doc_cache SET hits=hits+1 WHERE doc_hash=%s", (dh,))
            conn.commit()
            doc_summaries.append({"title": title, "summary": crow[0], "cached": True})
            continue

        full_path = os.path.join(docs_base, doc_path)
        tmp_path = None
        if not os.path.exists(full_path):
            tmp_path = _sftp_fetch(doc_path)            # prod: pull from ON storage over SFTP
            if tmp_path:
                full_path = tmp_path
        if not os.path.exists(full_path):
            doc_summaries.append({"title": title, "summary": "(File not found)"})
            continue
        low = doc_path.lower()                           # match on the real name, not the temp suffix
        if low.endswith(".pdf"):
            summary, usage = _summarise_pdf(full_path)
        elif low.endswith((".xlsx", ".xls")):
            summary, usage = _summarise_excel(full_path)
        else:
            summary, usage = f"(Unsupported file type: {os.path.splitext(doc_path)[1]})", {}
        if tmp_path:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        pt, ct = usage.get("prompt_tokens", 0), usage.get("completion_tokens", 0)
        ptok += pt
        ctok += ct
        if not summary.startswith("("):                 # cache only real summaries, not errors
            with conn.cursor() as cur:
                cur.execute("INSERT INTO buyer_match.doc_cache "
                            "(doc_hash, doc_path, title, summary, prompt_tokens, completion_tokens) "
                            "VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (doc_hash) DO NOTHING",
                            (dh, doc_path, title, summary, pt, ct))
            conn.commit()
        doc_summaries.append({"title": title, "summary": summary})

    return {
        "id": m["id"], "code": m.get("code", ""), "project_name": m.get("project_name", ""),
        "summary": m.get("summary", ""), "table": m["source_table"],
        "mandate_text": mandate_text, "doc_summaries": doc_summaries,
        "gpt_cost": {"prompt_tokens": ptok, "completion_tokens": ctok,
                     "total_tokens": ptok + ctok,
                     "cost_usd": (ptok * 0.15 + ctok * 0.60) / 1_000_000},
    }


def full_text(mandate_text, doc_summaries):
    """EXACT parity with the frontend fullText assembly (this is what gets embedded)."""
    ft = mandate_text
    if doc_summaries:
        ft += "\n\nDocument summaries:\n"
        for d in doc_summaries:
            ft += f"\n{d['title']}:\n{d['summary']}\n"
    return ft
